"""Receivables Ledger (MRI 'Rec Ledger') -> Excel.

The report interleaves TWO different column layouts on every page, which is why
a generic table detector shreds it (auto-detected column counts came out as
26/24/25 on three consecutive pages). So this reads the two block types by
their own fixed x-bands instead of trying to force one grid over the page:

  tenant line 1:  Suite | Tenant Name | Status | Contact
  tenant line 2:  Lease# | Lease Term: <from> To <to> | Square Feet: <n>
  summary rows:   CODE | Description | Mo.Rep | Beg Bal | CHARGES | Receipts ...

From the summary we keep only the charge code (+ its description) and the
Charges column. Those rows sit under their tenant with the identity columns
left blank, so the tenant reads as one block.
"""
from __future__ import annotations

import io
import logging
import os
import re
import tempfile

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# --- fixed x-bands of the summary block (PDF points) ---
_SUM_CODE_X = 92        # charge code sits left of this; description right of it
_SUM_DESC_END = 200     # description ends before the first money column
# The summary's money columns are right-aligned; these are their right edges,
# in printed order, matching the header row:
#   Mo. Rep Charges | Beg Balance | Charges | Cash Receipts
#   | N/C Credits | Refunds | End Balance | Sec Dep Bal
_SUM_MONEY_RIGHT = [272, 339, 402, 465, 528, 591, 654, 717]
_MONEY_FIELDS = ["mo_rep", "beg_bal", "charges", "receipts",
                 "nc_credits", "refunds", "end_bal", "sec_dep"]
_MONEY_HEADERS = ["Mo. Rep Charges", "Beg Balance", "Charges", "Cash Receipts",
                  "N/C Credits", "Refunds", "End Balance", "Sec Dep Bal"]
_CHARGES_RIGHT = _SUM_MONEY_RIGHT[2]   # used to detect a summary row

_DATE_RE = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}")
_NUM_RE = re.compile(r"^\(?-?[\d,]+\.\d{2}\)?$")
_CODE_RE = re.compile(r"^[A-Z][A-Z0-9&/]{0,5}$")


def _num(s: str):
    s = (s or "").strip()
    if not _NUM_RE.match(s):
        return None
    neg = s.startswith("(") and s.endswith(")")
    try:
        v = float(s.strip("()").replace(",", ""))
    except ValueError:
        return None
    return -v if neg else v


def _lines(page):
    words = page.extract_words(x_tolerance=1.5, y_tolerance=2) or []
    rows: dict[int, list] = {}
    for w in words:
        rows.setdefault(round(w["top"]), []).append(w)
    return [(t, sorted(rows[t], key=lambda w: w["x0"])) for t in sorted(rows)]


def _between(line, x0, x1):
    return " ".join(w["text"] for w in line if x0 <= w["x0"] < x1).strip()


def _money_at(line, right, tol=14):
    """The money figure whose RIGHT edge lands on this column."""
    for w in line:
        if abs(w["x1"] - right) <= tol:
            v = _num(w["text"])
            if v is not None:
                return v
    return None


def _after_label(text, label):
    i = text.find(label)
    return text[i + len(label):].strip() if i >= 0 else ""


class Row:
    __slots__ = ("suite", "lease", "tenant", "status", "contact",
                 "term_from", "term_to", "sqft", "code", "desc", "page",
                 *_MONEY_FIELDS)

    def __init__(self, **kw):
        for s in self.__slots__:
            setattr(self, s, kw.get(s, ""))


def _money_row(line) -> dict:
    """Every money column on a summary line, keyed by field name. A column the
    row doesn't print (charge rows carry no Sec Dep Bal) stays blank."""
    out = {}
    for field, right in zip(_MONEY_FIELDS, _SUM_MONEY_RIGHT):
        v = _money_at(line, right)
        if v is not None:
            out[field] = v
    return out


def parse(pdf_path: str) -> list[Row]:
    out: list[Row] = []
    pending: Row | None = None      # tenant awaiting its second line

    with pdfplumber.open(pdf_path) as pdf:
        for pno, page in enumerate(pdf.pages, start=1):
            for _y, line in _lines(page):
                text = " ".join(w["text"] for w in line)
                first = line[0]["text"]
                x0 = line[0]["x0"]

                # ---- summary row: a short CODE in the far-left band ----
                if x0 < _SUM_CODE_X and _CODE_RE.match(first) and len(line) > 2:
                    money = _money_row(line)
                    if "charges" in money:
                        out.append(Row(
                            page=pno, code=first,
                            desc=_between(line, _SUM_CODE_X, _SUM_DESC_END),
                            **money,
                        ))
                        continue

                # ---- summary total row ----
                if first == "Total:":
                    money = _money_row(line)
                    if money:
                        out.append(Row(page=pno, code="Total:", **money))
                    continue

                # ---- tenant line 2: 'Lease Term:' + 'Square Feet:' ----
                if "Lease Term:" in text and pending is not None:
                    pending.lease = first
                    seg = _after_label(text, "Lease Term:")
                    dates = _DATE_RE.findall(seg)
                    if dates:
                        pending.term_from = dates[0]
                        if len(dates) > 1:
                            pending.term_to = dates[1]
                    sq = _after_label(text, "Square Feet:")
                    if sq:
                        pending.sqft = sq.split()[0]
                    out.append(pending)
                    pending = None
                    continue

                # ---- tenant line 1: suite id then the name ----
                # Skip the column-header line and page furniture.
                if (x0 < _SUM_CODE_X and len(line) >= 3
                        and first not in ("Lease", "Suite", "Category", "Date")
                        and not _NUM_RE.match(first)
                        and not _DATE_RE.match(first)):
                    name = _between(line, 140, 300)
                    if name:
                        pending = Row(page=pno, suite=first, tenant=name,
                                      status=_between(line, 300, 440),
                                      contact=_between(line, 440, 620))
                    continue

    return out


HEADERS = (["Page", "Suite", "Lease", "Tenant Name", "Status", "Contact",
            "Term From", "Term To", "Square Feet", "Charge Code",
            "Charge Description"] + _MONEY_HEADERS)
_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")


def build_xlsx(rows: list[Row]) -> tuple[bytes, int]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = _FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 2
    for row in rows:
        ws.cell(row=r, column=1, value=row.page)
        # Identity columns stay blank on charge rows — they belong to the
        # tenant line above, so each tenant reads as one block.
        ws.cell(row=r, column=2, value=row.suite)
        ws.cell(row=r, column=3, value=row.lease)
        ws.cell(row=r, column=4, value=row.tenant)
        ws.cell(row=r, column=5, value=row.status)
        ws.cell(row=r, column=6, value=row.contact)
        ws.cell(row=r, column=7, value=row.term_from)
        ws.cell(row=r, column=8, value=row.term_to)
        ws.cell(row=r, column=9, value=row.sqft)
        ws.cell(row=r, column=10, value=row.code)
        ws.cell(row=r, column=11, value=row.desc)
        bold = row.code == "Total:"
        for j, field in enumerate(_MONEY_FIELDS):
            v = getattr(row, field)
            if v == "":
                continue
            c = ws.cell(row=r, column=12 + j, value=v)
            c.number_format = "#,##0.00"
            if bold:
                c.font = Font(bold=True)
        if bold:
            ws.cell(row=r, column=10).font = Font(bold=True)
        r += 1

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        longest = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 42)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), r - 2


def convert(pdf_bytes: bytes) -> tuple[bytes, int]:
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        tmp = f.name
    try:
        return build_xlsx(parse(tmp))
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass
