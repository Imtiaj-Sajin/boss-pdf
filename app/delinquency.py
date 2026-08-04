"""MRI 'Aged Delinquencies Report' (Report Id MRIX_CMAGEDL) -> Excel.

One PDF = one property/building. A folder of them combines into a single sheet
that stacks every property, each ending in its own 'Grand Total:' — matching the
analyst's 'Working.xls' layout exactly.

Key quirk: the property name is NOT in the PDF body. It lives only in the file
name (e.g. '1598_B4__595_South_Broadway__Aged_Delinquency_Report...pdf' ->
'1598_B4 - 595 South Broadway'), so we reconstruct it from there.
"""
from __future__ import annotations

import io
import logging
import os
import re
from dataclasses import dataclass, field

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
_INVOICE_RE = re.compile(r"^[A-Za-z0-9]+-[A-Za-z0-9]+$")
_NUM_RE = re.compile(r"^\(?-?\$?[\d,]+(?:\.\d+)?\)?$")

# x-bands (PDF points) of the charge-line columns, from the report's fixed layout
_X_CODE = 85     # Category code (HT/GS/SB/RT/BDR) starts here
_X_DESC = 116    # Category Type description
_X_SRC = 222     # Src. (CH)
_X_NUM = 258     # the six money figures start to the right of this


@dataclass
class Charge:
    date: str = ""
    category: str = ""        # short code
    category_type: str = ""   # description
    src: str = ""
    amount: float = 0.0
    current: float = 0.0
    d30: float = 0.0
    d60: float = 0.0
    d90: float = 0.0
    d120: float = 0.0


@dataclass
class Tenant:
    invoice_no: str = ""
    name: str = ""
    suite: str = ""
    master_id: str = ""
    status: str = ""
    charges: list[Charge] = field(default_factory=list)
    total: list[float] | None = None   # [amount, current, 30, 60, 90, 120]


@dataclass
class Property:
    name: str = ""
    as_of: str = ""
    tenants: list[Tenant] = field(default_factory=list)
    grand_total: list[float] | None = None


# ---------- helpers ----------

def _num(s: str):
    s = (s or "").strip()
    if not s or not _NUM_RE.match(s):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def property_name_from_filename(filename: str) -> str:
    """'1598_B4__595_South_Broadway__Aged_...' -> '1598_B4 - 595 South Broadway'."""
    base = os.path.splitext(os.path.basename(filename or ""))[0]
    parts = [p for p in base.split("__") if p]
    segs = []
    for p in parts:
        if p.lower().startswith("aged"):
            break
        segs.append(p)
    if not segs:
        return base or "Property"
    if len(segs) >= 2:
        code = segs[0]
        addr = " ".join(s.replace("_", " ") for s in segs[1:]).strip()
        return f"{code} - {addr}"
    return segs[0].replace("_", " ")


def _lines(page):
    """Words grouped into visual lines, each sorted left-to-right."""
    words = page.extract_words(x_tolerance=1.5, y_tolerance=2) or []
    rows: dict[int, list] = {}
    for w in words:
        rows.setdefault(round(w["top"]), []).append(w)
    return [sorted(rows[t], key=lambda w: w["x0"]) for t in sorted(rows)]


def _text(line) -> str:
    return " ".join(w["text"] for w in line)


# ---------- parse ----------

def parse(pdf_path: str, filename: str) -> Property:
    prop = Property(name=property_name_from_filename(filename))
    cur: Tenant | None = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for line in _lines(page):
                t = _text(line)
                first = line[0]["text"] if line else ""

                if prop.as_of == "" and t.startswith("Report Date:"):
                    m = _DATE_RE.pattern  # noqa
                    mm = re.search(r"\d{1,2}/\d{1,2}/\d{4}", t)
                    if mm:
                        prop.as_of = mm.group(0)
                    continue

                if "Grand Total:" in t:
                    prop.grand_total = _trailing_numbers(line)
                    cur = None
                    continue

                # Tenant total: '<name> Total:' with the six figures (not Grand).
                # The total line carries the CLEAN tenant name (the header line
                # can fuse it with 'Suite' — e.g. 'HOLDINGS, LLSCuite'), so use
                # it as the authoritative name.
                if "Total:" in t and not _DATE_RE.match(first):
                    if cur is not None:
                        cur.total = _trailing_numbers(line)
                        clean = t.split(" Total:")[0].strip()
                        if clean:
                            cur.name = clean
                    cur = None
                    continue

                # Tenant header: an invoice code followed by 'Day Due:'. We key on
                # 'Day Due:' (present on every header, nowhere else) rather than the
                # word 'Suite', which the PDF sometimes glues to the tenant name.
                if _INVOICE_RE.match(first) and (
                        "Day Due:" in t or ("Suite" in t and "Id:" in t)):
                    cur = _parse_tenant_header(line)
                    prop.tenants.append(cur)
                    continue

                if "Master Occupant Id:" in t and cur is not None:
                    cur.master_id = _after(line, "Id:")
                    continue

                if first == "Contact:" and cur is not None:
                    cur.status = _after(line, "Status:")
                    continue

                # Charge line: starts with a date
                if _DATE_RE.match(first):
                    ch = _parse_charge(line)
                    if ch is not None and cur is not None:
                        cur.charges.append(ch)
                    continue

    return prop


def _trailing_numbers(line) -> list[float]:
    nums = [_num(w["text"]) for w in line if _num(w["text"]) is not None]
    nums = nums[-6:]
    while len(nums) < 6:
        nums.append(0.0)
    return nums


def _after(line, token: str) -> str:
    texts = [w["text"] for w in line]
    if token in texts:
        i = texts.index(token)
        if i + 1 < len(texts):
            return texts[i + 1]
    return ""


def _parse_tenant_header(line) -> Tenant:
    texts = [w["text"] for w in line]
    inv = texts[0]
    # Name runs from after the invoice up to the 'Suite' marker — which may be a
    # standalone token OR fused onto the previous word ('...LLSCuite'). Stop at
    # the first token that is/contains that marker, or at 'Id:'. (The Total line
    # later overrides this with the clean name anyway.)
    stop = len(texts)
    for j in range(1, len(texts)):
        tk = texts[j]
        if tk == "Id:" or tk == "Suite" or tk.endswith("uite"):
            stop = j
            break
    name = " ".join(texts[1:stop]).strip().rstrip(",")
    # suite = the token right after the first 'Id:'
    suite = ""
    if "Id:" in texts:
        i = texts.index("Id:")
        if i + 1 < len(texts):
            suite = texts[i + 1]
    return Tenant(invoice_no=inv, name=name, suite=suite)


def _parse_charge(line) -> Charge | None:
    date = line[0]["text"]
    code, desc, src = [], [], []
    nums = []
    for w in line[1:]:
        x, txt = w["x0"], w["text"]
        n = _num(txt)
        if n is not None and x >= _X_NUM:
            nums.append((x, n))
        elif x < _X_CODE:
            continue
        elif x < _X_DESC:
            code.append(txt)
        elif x < _X_SRC:
            desc.append(txt)
        elif x < _X_NUM:
            src.append(txt)
    nums = [n for _, n in sorted(nums)][-6:]
    while len(nums) < 6:
        nums.append(0.0)
    return Charge(
        date=date, category=" ".join(code), category_type=" ".join(desc),
        src=" ".join(src), amount=nums[0], current=nums[1],
        d30=nums[2], d60=nums[3], d90=nums[4], d120=nums[5],
    )


# ---------- Excel (matches Working.xls) ----------

HEADERS = ["Invoice Date", "Invoice No", "Property Name", "Tenant Name",
           "Suite Id:", "Master Occupant Id:", "Status", "Category",
           "Category Type", "Src.", "Amount", "Current", 30, 60, 90, 120]
COL0 = 2  # first data column is B
_HEADER_FONT = Font(bold=True)


def _money_cols(ws, row, vals):
    """Write [amount, current, 30, 60, 90, 120] into L..Q."""
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=12 + i)   # L=12
        c.value = int(v) if float(v).is_integer() else round(v, 2)


def build_combined_xlsx(properties: list[Property]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "All pages"

    # Row 1 is intentionally left blank (matches the analyst's Working.xls):
    # As-of on row 2, headers on row 3, data from row 4.
    as_of = next((p.as_of for p in properties if p.as_of), "")
    ws.cell(row=2, column=COL0, value=f"As of {as_of}".rstrip())
    for j, h in enumerate(HEADERS):
        ws.cell(row=3, column=COL0 + j, value=h).font = _HEADER_FONT

    r = 4
    for prop in properties:
        first_row_of_prop = True
        for tn in prop.tenants:
            if not tn.charges and tn.total is None:
                continue
            first_charge = True
            for ch in tn.charges:
                ws.cell(row=r, column=2, value=ch.date)          # B Invoice Date
                ws.cell(row=r, column=3, value=tn.invoice_no)    # C Invoice No
                if first_charge:
                    if first_row_of_prop:
                        ws.cell(row=r, column=4, value=prop.name)      # D Property
                        first_row_of_prop = False
                    ws.cell(row=r, column=5, value=tn.name)            # E Tenant
                    ws.cell(row=r, column=6,
                            value=int(tn.suite) if tn.suite.isdigit() else tn.suite)  # F Suite
                    ws.cell(row=r, column=7, value=tn.master_id)       # G Master Occ.
                    ws.cell(row=r, column=8, value=tn.status)          # H Status
                    first_charge = False
                ws.cell(row=r, column=9, value=ch.category)            # I Category
                ws.cell(row=r, column=10, value=ch.category_type)      # J Category Type
                ws.cell(row=r, column=11, value=ch.src)                # K Src.
                _money_cols(ws, r, [ch.amount, ch.current, ch.d30, ch.d60, ch.d90, ch.d120])
                r += 1
            if tn.total is not None:
                ws.cell(row=r, column=10, value=f"{tn.name} Total:")   # J label
                _money_cols(ws, r, tn.total)
                r += 1
        # blank row, then this property's Grand Total
        r += 1
        if prop.grand_total is not None:
            ws.cell(row=r, column=10, value="Grand Total:")
            _money_cols(ws, r, prop.grand_total)
            r += 1

    _autosize(ws)
    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 40)
