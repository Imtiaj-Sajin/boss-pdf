"""Data Puller — pull named rectangular regions from many same-layout PDFs into
one row-per-file Excel table.

Each region is given in **normalized** coordinates (fractions of the page, top-
left origin), so a selector drawn once on a reference PDF applies to every other
PDF with the same layout. Per region we read the PDF's native text layer inside
the box; if that box has no text (a scanned page), we fall back to OCR on a
rendered crop. Reuses the main converter's renderer + OCR engines.
"""
from __future__ import annotations

import io
import logging
import os
import re
import tempfile
from dataclasses import dataclass

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import converter as _conv

logger = logging.getLogger(__name__)


@dataclass
class Region:
    name: str
    page: int          # 1-indexed
    x0: float          # all four are normalized [0,1], top-left origin
    y0: float
    x1: float
    y1: float


def _clip01(v) -> float:
    try:
        return max(0.0, min(1.0, float(v)))
    except (TypeError, ValueError):
        return 0.0


def _clean(s: str) -> str:
    if not s:
        return ""
    # collapse internal whitespace/newlines into single spaces
    return re.sub(r"\s+", " ", str(s)).strip()


def _bbox_for(region: Region, width: float, height: float):
    """Normalized region -> absolute (x0, top, x1, bottom), normalized so x0<x1."""
    x0 = _clip01(region.x0) * width
    x1 = _clip01(region.x1) * width
    top = _clip01(region.y0) * height
    bottom = _clip01(region.y1) * height
    if x1 < x0:
        x0, x1 = x1, x0
    if bottom < top:
        top, bottom = bottom, top
    return x0, top, x1, bottom


def _native_text(page, region: Region) -> str:
    x0, top, x1, bottom = _bbox_for(region, float(page.width), float(page.height))
    if x1 - x0 < 1 or bottom - top < 1:
        return ""
    try:
        crop = page.within_bbox((x0, top, x1, bottom))
        return _clean(crop.extract_text(x_tolerance=2, y_tolerance=2) or "")
    except Exception as e:
        logger.debug("native region extract failed: %s", e)
        return ""


def _ocr_image_text(img) -> str:
    """Best-effort OCR of a small cropped region image. Tries Tesseract (simple
    line read) then RapidOCR. Returns '' if no OCR engine is available."""
    if _conv._HAS_TESSERACT:
        try:
            import pytesseract
            txt = pytesseract.image_to_string(img, config="--psm 6")
            if txt and txt.strip():
                return _clean(txt)
        except Exception as e:
            logger.debug("tesseract region OCR failed: %s", e)
    engine = _conv._get_rapidocr()
    if engine is not None:
        try:
            import numpy as np
            arr = np.array(img.convert("RGB"))
            result, _ = engine(arr)
            if result:
                return _clean(" ".join((e[1] or "") for e in result))
        except Exception as e:
            logger.debug("rapidocr region OCR failed: %s", e)
    return ""


def _ocr_text(pdf_path: str, region: Region) -> str:
    """Render the region's page, crop to the box, OCR it."""
    if not _conv._HAS_FITZ:
        return ""
    img = _conv._render_page_image(pdf_path, region.page, dpi=300)
    if img is None:
        return ""
    W, H = img.size
    x0 = int(_clip01(region.x0) * W)
    x1 = int(_clip01(region.x1) * W)
    y0 = int(_clip01(region.y0) * H)
    y1 = int(_clip01(region.y1) * H)
    if x1 < x0:
        x0, x1 = x1, x0
    if y1 < y0:
        y0, y1 = y1, y0
    if x1 - x0 < 2 or y1 - y0 < 2:
        return ""
    try:
        crop = img.crop((x0, y0, x1, y1))
    except Exception:
        return ""
    return _ocr_image_text(crop)


def extract(pdf_bytes: bytes, regions: list[Region]) -> dict[str, str]:
    """Pull every region's value from one PDF. Native text first; OCR fallback
    when a box yields nothing. Returns {region_name: value}."""
    out: dict[str, str] = {}
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        tmp = f.name
    try:
        with pdfplumber.open(tmp) as pdf:
            n = len(pdf.pages)
            for r in regions:
                val = ""
                if 1 <= r.page <= n:
                    val = _native_text(pdf.pages[r.page - 1], r)
                # OCR fallback (only meaningful if engines are installed)
                if not val and 1 <= r.page <= n:
                    val = _ocr_text(tmp, r)
                # last region with the same name wins only if it found text;
                # keep the first non-empty value otherwise.
                if r.name in out and out[r.name] and not val:
                    continue
                out[r.name] = val
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass
    return out


# ---------- Excel: one row per PDF, one column per datapoint ----------

_HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
_THIN = Side(border_style="thin", color="FFD9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_xlsx(filenames: list[str], rows: list[dict[str, str]],
               column_names: list[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Puller"

    headers = ["File"] + column_names
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = _HEADER_FILL
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = _BORDER

    for i, (fn, row) in enumerate(zip(filenames, rows), start=2):
        fc = ws.cell(row=i, column=1, value=fn)
        fc.border = _BORDER
        fc.alignment = Alignment(vertical="center")
        for j, name in enumerate(column_names, start=2):
            cc = ws.cell(row=i, column=j, value=row.get(name, ""))
            cc.border = _BORDER
            cc.alignment = Alignment(vertical="center", wrap_text=True)

    # autosize columns (cap width)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            for line in str(v).splitlines():
                longest = max(longest, len(line))
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 50)

    ws.freeze_panes = "B2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
