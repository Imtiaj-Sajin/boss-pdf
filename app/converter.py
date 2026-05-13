"""PDF -> Excel conversion engine.

Per-page best-of strategy:
  1. pdfplumber           — native PDFs with a text layer (preserves styling)
  2. Camelot (lattice)    — tables with visible borders
  3. Camelot (stream)     — tables without borders
  4. PaddleOCR PP-Structure — scanned pages, layout + table-aware OCR
  5. Tesseract (smart)    — final fallback, with row/column clustering

Pages are auto-detected as scanned (no text layer) and routed through OCR.
Page rendering uses PyMuPDF, so Poppler is NOT required.
"""
from __future__ import annotations

import io
import logging
import os
import re
import tempfile
from collections import Counter
from dataclasses import dataclass, field
from html.parser import HTMLParser
from typing import Iterable, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---- Optional deps (imported lazily / behind capability flags) ----

try:
    import camelot  # type: ignore
    _HAS_CAMELOT = True
except Exception as e:
    logger.warning("camelot unavailable: %s", e)
    _HAS_CAMELOT = False

# PyMuPDF: page -> image rendering without Poppler.
try:
    import fitz  # PyMuPDF
    from PIL import Image  # noqa: F401
    _HAS_FITZ = True
except Exception as e:
    logger.warning("PyMuPDF unavailable: %s", e)
    _HAS_FITZ = False

# PaddleOCR PP-Structure: layout + table-aware OCR. Heavy first import.
try:
    import paddleocr  # noqa: F401
    _HAS_PADDLE = True
except Exception as e:
    logger.info("paddleocr not available: %s", e)
    _HAS_PADDLE = False

_PADDLE_ENGINE = None  # cached engine (or False if init failed)
_PADDLE_API_VERSION = 0  # 0=none, 2=PPStructure (legacy), 3=PPStructureV3

# RapidOCR: ONNX-based port of PaddleOCR. Lighter, broader Python compat,
# no paddlepaddle dep. Used as primary OCR engine when paddle isn't available.
try:
    import rapidocr_onnxruntime  # noqa: F401
    _HAS_RAPIDOCR = True
except Exception as e:
    logger.info("rapidocr_onnxruntime not available: %s", e)
    _HAS_RAPIDOCR = False

_RAPIDOCR_ENGINE = None  # cached RapidOCR instance (or False if init failed)

# Tesseract OCR (fallback only; needs system Tesseract binary).
try:
    import pytesseract  # type: ignore
    _HAS_TESSERACT = True
except Exception as e:
    logger.info("tesseract OCR not available: %s", e)
    _HAS_TESSERACT = False

# pdf2image is a last-resort renderer (needs Poppler).
try:
    from pdf2image import convert_from_path  # type: ignore
    _HAS_PDF2IMAGE = True
except Exception:
    _HAS_PDF2IMAGE = False


# ---------- Rich data model ----------

@dataclass
class CellStyle:
    bold: bool = False
    italic: bool = False
    color_hex: Optional[str] = None  # 6-digit hex, no alpha, e.g. "FF0000"


@dataclass
class RichCell:
    text: str = ""
    style: CellStyle = field(default_factory=CellStyle)


RichTable = list[list[RichCell]]


@dataclass
class PageResult:
    page_num: int  # 1-indexed
    tables: list[RichTable]
    title_lines: list[str]
    method: str
    score: float
    char_count: int = 0  # number of native-text chars on the page; 0 → likely scanned


# ---------- Char-level helpers (pdfplumber) ----------

def _chars_in_bbox(chars: list[dict], bbox: tuple[float, float, float, float]) -> list[dict]:
    x0, top, x1, bottom = bbox
    out = []
    for c in chars:
        cx = (c["x0"] + c["x1"]) / 2
        cy = (c["top"] + c["bottom"]) / 2
        if x0 <= cx <= x1 and top <= cy <= bottom:
            out.append(c)
    return out


def _is_bold(chars: list[dict]) -> bool:
    if not chars:
        return False
    bold_kw = ("bold", "black", "heavy", "semibold", "demibold")
    n_bold = sum(1 for c in chars
                 if any(k in (c.get("fontname") or "").lower() for k in bold_kw))
    return n_bold > len(chars) / 2


def _color_to_hex(color) -> Optional[str]:
    """Convert a pdfplumber non_stroking_color value to '#RRGGBB' (no '#')."""
    if color is None:
        return None
    if isinstance(color, (int, float)):
        v = max(0, min(255, int(round(float(color) * 255))))
        return f"{v:02X}{v:02X}{v:02X}"
    if isinstance(color, (list, tuple)):
        nums = [float(x) for x in color if isinstance(x, (int, float))]
        if len(nums) == 1:
            v = max(0, min(255, int(round(nums[0] * 255))))
            return f"{v:02X}{v:02X}{v:02X}"
        if len(nums) == 3:
            r, g, b = (max(0, min(255, int(round(x * 255)))) for x in nums)
            return f"{r:02X}{g:02X}{b:02X}"
        if len(nums) == 4:
            c, m, y, k = nums
            r = (1 - c) * (1 - k)
            g = (1 - m) * (1 - k)
            b = (1 - y) * (1 - k)
            r, g, b = (max(0, min(255, int(round(v * 255)))) for v in (r, g, b))
            return f"{r:02X}{g:02X}{b:02X}"
    return None


def _dominant_color(chars: list[dict]) -> Optional[str]:
    if not chars:
        return None
    counts: Counter = Counter()
    for c in chars:
        col = c.get("non_stroking_color")
        hexv = _color_to_hex(col)
        if hexv and hexv != "000000":
            counts[hexv] += 1
    return counts.most_common(1)[0][0] if counts else None


# ---------- Text/cell normalization ----------

def _clean_cell_text(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = str(s).replace(" ", " ").strip()
    return "\n".join(" ".join(line.split()) for line in s.splitlines())


def _normalize_rich_table(rows: list[list[RichCell]]) -> RichTable:
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    for r in rows:
        if len(r) < width:
            r.extend([RichCell() for _ in range(width - len(r))])
    while rows and all(c.text == "" for c in rows[-1]):
        rows.pop()
    while rows and all(c.text == "" for c in rows[0]):
        rows.pop(0)
    return rows


def _normalize_plain_table(rows: Iterable[Iterable]) -> RichTable:
    return _normalize_rich_table(
        [[RichCell(text=_clean_cell_text(c)) for c in r] for r in rows]
    )


def _score_table(t: RichTable) -> float:
    if not t or not t[0]:
        return 0.0
    cells = sum(len(r) for r in t)
    if cells == 0:
        return 0.0
    filled = sum(1 for row in t for c in row if c.text)
    return filled * (0.5 + 0.5 * (filled / cells))


# ---------- Extraction: pdfplumber with rich style ----------

_TABLE_SETTINGS_OPTIONS = (
    {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
    {"vertical_strategy": "lines", "horizontal_strategy": "text",
     "intersection_tolerance": 5},
    {"vertical_strategy": "text", "horizontal_strategy": "text",
     "intersection_tolerance": 5},
)


def _extract_rich_table_from_pdfplumber(page, table) -> RichTable:
    """Pull a pdfplumber Table into a RichTable with style info from chars."""
    chars = page.chars
    rows: list[list[RichCell]] = []
    for row in table.rows:
        rich_row: list[RichCell] = []
        for cell_bbox in row.cells:
            if cell_bbox is None:
                rich_row.append(RichCell())
                continue
            try:
                sub = page.within_bbox(cell_bbox)
                text = _clean_cell_text(sub.extract_text(x_tolerance=2, y_tolerance=2))
            except Exception:
                text = ""
            cell_chars = _chars_in_bbox(chars, cell_bbox)
            style = CellStyle(
                bold=_is_bold(cell_chars),
                color_hex=_dominant_color(cell_chars),
            )
            rich_row.append(RichCell(text=text, style=style))
        rows.append(rich_row)
    return _normalize_rich_table(rows)


def _extract_page_title_lines(page, table_bboxes: list[tuple]) -> list[str]:
    """Lines of text on the page above the first table, keep their original order."""
    if not table_bboxes:
        return []
    top_of_tables = min(b[1] for b in table_bboxes)
    if top_of_tables <= 5:
        return []
    try:
        crop = page.within_bbox((0, 0, float(page.width), float(top_of_tables)))
        text = crop.extract_text(x_tolerance=2, y_tolerance=2) or ""
    except Exception:
        return []
    lines = [ln.strip() for ln in text.splitlines()]
    return [ln for ln in lines if ln]


def _extract_pdfplumber(pdf_path: str, pages: Optional[set[int]] = None) -> list[PageResult]:
    results: list[PageResult] = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            if pages is not None and i not in pages:
                continue
            tables: list[RichTable] = []
            table_bboxes: list[tuple] = []
            for settings in _TABLE_SETTINGS_OPTIONS:
                try:
                    found = page.find_tables(table_settings=settings) or []
                except Exception as e:
                    logger.warning("find_tables p%d %s: %s", i, settings, e)
                    found = []
                for t in found:
                    try:
                        rt = _extract_rich_table_from_pdfplumber(page, t)
                        if rt:
                            tables.append(rt)
                            table_bboxes.append(t.bbox)
                    except Exception as e:
                        logger.warning("rich extract p%d failed: %s", i, e)
                if tables:
                    break

            title_lines = _extract_page_title_lines(page, table_bboxes)
            score = sum(_score_table(t) for t in tables)
            try:
                char_count = len(page.chars)
            except Exception:
                char_count = 0
            results.append(PageResult(
                page_num=i, tables=tables, title_lines=title_lines,
                method="pdfplumber", score=score, char_count=char_count,
            ))
    return results


# ---------- Extraction: camelot (fallback, plain text) ----------

def _camelot_pages(pdf_path: str, flavor: str, page_count: int,
                   pages: Optional[set[int]] = None) -> dict[int, list[RichTable]]:
    out: dict[int, list[RichTable]] = {}
    if not _HAS_CAMELOT:
        return out
    if pages is not None:
        if not pages:
            return out
        page_spec = ",".join(str(p) for p in sorted(pages))
    else:
        page_spec = f"1-{page_count}"
    try:
        tlist = camelot.read_pdf(pdf_path, pages=page_spec,
                                 flavor=flavor, suppress_stdout=True)
    except Exception as e:
        logger.warning("camelot %s failed: %s", flavor, e)
        return out
    for t in tlist:
        try:
            page = int(t.page)
            rows = t.df.values.tolist()
            norm = _normalize_plain_table(rows)
            if norm:
                out.setdefault(page, []).append(norm)
        except Exception as e:
            logger.warning("camelot %s parse failed: %s", flavor, e)
    return out


# ---------- Page rendering (PyMuPDF first, Poppler last-resort) ----------

def _render_page_image(pdf_path: str, page_num: int, dpi: int = 300):
    """Render a single PDF page to a PIL.Image. Prefers PyMuPDF (no Poppler)."""
    if _HAS_FITZ:
        try:
            doc = fitz.open(pdf_path)
            try:
                page = doc[page_num - 1]
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                mode = "RGB" if pix.n in (3, 4) else "L"
                img = Image.frombytes(mode, (pix.width, pix.height), pix.samples)
                if img.mode != "RGB":
                    img = img.convert("RGB")
                return img
            finally:
                doc.close()
        except Exception as e:
            logger.warning("PyMuPDF render p%d failed: %s", page_num, e)
    if _HAS_PDF2IMAGE:
        try:
            imgs = convert_from_path(pdf_path, dpi=dpi,
                                     first_page=page_num, last_page=page_num)
            return imgs[0] if imgs else None
        except Exception as e:
            logger.warning("pdf2image render p%d failed: %s", page_num, e)
    return None


# ---------- HTML table parser (used for PP-Structure output) ----------

class _HtmlTableParser(HTMLParser):
    """Parse a single <table>...</table> string into a RichTable, honoring
    <th> bold styling and basic colspan/rowspan."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._rows: list[list[Optional[RichCell]]] = []
        self._row: Optional[list[Optional[RichCell]]] = None
        self._cell_text: Optional[list[str]] = None
        self._cell_is_th = False
        self._cell_colspan = 1
        self._cell_rowspan = 1
        self._bold_depth = 0
        # rowspan tracking: cells "owed" to upcoming rows by column index
        self._rowspan_owed: list[int] = []

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "tr":
            self._row = []
            # apply any rowspans owed from earlier rows
            for ci, owed in enumerate(self._rowspan_owed):
                if owed > 0:
                    while len(self._row) <= ci:
                        self._row.append(None)
                    self._row[ci] = RichCell()  # blank merged-cell placeholder
        elif tag in ("td", "th"):
            self._cell_text = []
            self._cell_is_th = (tag == "th")
            d = dict(attrs)
            try:
                self._cell_colspan = max(1, int(d.get("colspan", 1)))
            except ValueError:
                self._cell_colspan = 1
            try:
                self._cell_rowspan = max(1, int(d.get("rowspan", 1)))
            except ValueError:
                self._cell_rowspan = 1
        elif tag in ("b", "strong"):
            self._bold_depth += 1
        elif tag == "br":
            if self._cell_text is not None:
                self._cell_text.append("\n")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag == "tr" and self._row is not None:
            self._rows.append(self._row)
            # decrement owed-rowspans
            self._rowspan_owed = [max(0, n - 1) for n in self._rowspan_owed]
            self._row = None
        elif tag in ("td", "th") and self._row is not None and self._cell_text is not None:
            text = _clean_cell_text(" ".join(self._cell_text))
            style = CellStyle(bold=self._cell_is_th or self._bold_depth > 0)
            cell = RichCell(text=text, style=style)
            # find first empty slot (None) at the end
            while self._row and self._row[-1] is None:
                pass  # keep as-is
            # append the real cell + colspan filler
            target_idx = len(self._row)
            # if this position was reserved by a rowspan, shift right
            while target_idx < len(self._row) and self._row[target_idx] is not None:
                target_idx += 1
            while len(self._row) < target_idx:
                self._row.append(None)
            self._row.append(cell)
            for _ in range(self._cell_colspan - 1):
                self._row.append(RichCell())
            # register rowspan so subsequent rows skip this column
            if self._cell_rowspan > 1:
                col_idx = len(self._row) - self._cell_colspan
                while len(self._rowspan_owed) <= col_idx:
                    self._rowspan_owed.append(0)
                self._rowspan_owed[col_idx] = max(
                    self._rowspan_owed[col_idx], self._cell_rowspan - 1
                )
            self._cell_text = None
            self._cell_is_th = False
            self._cell_colspan = 1
            self._cell_rowspan = 1
        elif tag in ("b", "strong"):
            self._bold_depth = max(0, self._bold_depth - 1)

    def handle_data(self, data):
        if self._cell_text is not None:
            self._cell_text.append(data)

    def result(self) -> list[list[RichCell]]:
        return [
            [(c if c is not None else RichCell()) for c in row]
            for row in self._rows
        ]


def _parse_html_table(html: str) -> Optional[RichTable]:
    if not html or "<" not in html:
        return None
    try:
        p = _HtmlTableParser()
        p.feed(html)
        p.close()
    except Exception as e:
        logger.warning("HTML table parse failed: %s", e)
        return None
    rows = p.result()
    return _normalize_rich_table(rows) if rows else None


# ---------- Extraction: PaddleOCR PP-Structure (best for scanned tables) ----------

def _get_paddle_engine():
    """Lazily build a PaddleOCR layout/table engine.

    Tries paddleocr 3.x (PPStructureV3) first — the current SOTA pipeline —
    then falls back to legacy PPStructure (paddleocr 2.x). Caches result.
    """
    global _PADDLE_ENGINE, _PADDLE_API_VERSION
    if _PADDLE_ENGINE is False:
        return None
    if _PADDLE_ENGINE is not None:
        return _PADDLE_ENGINE
    if not _HAS_PADDLE:
        _PADDLE_ENGINE = False
        return None

    # ---- paddleocr 3.x: PPStructureV3 (preferred) ----
    try:
        from paddleocr import PPStructureV3  # type: ignore
        logger.info("Initializing PPStructureV3 (paddleocr 3.x; downloads models on first use)…")
        try:
            _PADDLE_ENGINE = PPStructureV3(
                use_doc_orientation_classify=False,
                use_doc_unwarping=False,
                use_textline_orientation=False,
            )
        except TypeError:
            # Some 3.x point releases use slightly different kwarg names
            _PADDLE_ENGINE = PPStructureV3()
        _PADDLE_API_VERSION = 3
        logger.info("PPStructureV3 ready (paddleocr v3 API).")
        return _PADDLE_ENGINE
    except ImportError:
        logger.debug("PPStructureV3 not available; trying legacy PPStructure.")
    except Exception as e:
        logger.warning("PPStructureV3 init failed: %s", e)

    # ---- paddleocr 2.x: PPStructure (legacy fallback) ----
    try:
        from paddleocr import PPStructure  # type: ignore
        logger.info("Initializing PPStructure (paddleocr 2.x)…")
        _PADDLE_ENGINE = PPStructure(
            table=True, ocr=True, layout=True,
            show_log=False, lang="en", use_gpu=False,
        )
        _PADDLE_API_VERSION = 2
        logger.info("PPStructure ready (paddleocr v2 API).")
        return _PADDLE_ENGINE
    except Exception as e:
        logger.warning("PPStructure init failed: %s", e)

    _PADDLE_ENGINE = False
    return None


def _paddle_extract_page(pdf_path: str, page_num: int) -> tuple[list[RichTable], list[str]]:
    """Run paddleocr layout/table pipeline on one page. Handles both
    PPStructureV3 (v3.x, current SOTA) and legacy PPStructure (v2.x)."""
    engine = _get_paddle_engine()
    if engine is None:
        return [], []

    img = _render_page_image(pdf_path, page_num, dpi=300)
    if img is None:
        return [], []

    try:
        import numpy as np
        arr = np.array(img)
        if arr.ndim == 2:
            arr = np.stack([arr] * 3, axis=-1)
        elif arr.shape[-1] == 4:
            arr = arr[..., :3]
    except Exception as e:
        logger.warning("paddle prep p%d failed: %s", page_num, e)
        return [], []

    if _PADDLE_API_VERSION == 3:
        return _paddle_v3_extract(engine, arr, page_num)
    return _paddle_v2_extract(engine, arr, page_num)


def _paddle_v3_extract(engine, arr, page_num: int) -> tuple[list[RichTable], list[str]]:
    """PPStructureV3 result schema:
       result = [ {'table_res_list': [...], 'overall_ocr_res': {...}, ...} ]"""
    try:
        result = engine.predict(arr)
    except Exception as e:
        logger.warning("PPStructureV3 predict p%d failed: %s", page_num, e)
        return [], []
    if not result:
        return [], []

    # PPStructureV3 returns one result-object per input image. Single image -> [0].
    page = result[0] if isinstance(result, list) else result
    # 3.x output objects expose attribute-style access via .json/.dict but in
    # most builds also act as dicts. Read defensively.
    def _get(o, key, default=None):
        if isinstance(o, dict):
            return o.get(key, default)
        return getattr(o, key, default)

    tables: list[RichTable] = []
    titles: list[str] = []

    # ---- tables ----
    table_res_list = _get(page, "table_res_list") or []
    for tr in table_res_list:
        html = _get(tr, "pred_html") or _get(tr, "html") or ""
        rt = _parse_html_table(html)
        if rt:
            tables.append(rt)
            logger.debug("paddle v3 p%d: table parsed (%d rows × %d cols)",
                         page_num, len(rt), len(rt[0]) if rt else 0)
            continue
        # Fall back to cell-level reconstruction if html parse failed
        cells = _get(tr, "cell_box_list") or _get(tr, "cells") or []
        if cells:
            words = []
            for c in cells:
                txt = _get(c, "text") or _get(c, "rec_text") or ""
                txt = (txt or "").strip()
                if not txt:
                    continue
                box = _get(c, "box") or _get(c, "bbox")
                if box and len(box) >= 4:
                    try:
                        x0, y0, x1, y1 = (float(box[0]), float(box[1]),
                                          float(box[2]), float(box[3]))
                        words.append({
                            "text": txt,
                            "x": int(x0), "y": int(y0),
                            "w": max(1, int(x1 - x0)),
                            "h": max(1, int(y1 - y0)),
                        })
                    except (TypeError, ValueError):
                        pass
            if words:
                rt = _build_table_from_words(words)
                if rt:
                    tables.append(rt)

    # ---- standalone OCR text (only used if no tables found) ----
    if not tables:
        ocr_res = _get(page, "overall_ocr_res") or {}
        rec_texts = _get(ocr_res, "rec_texts") or []
        rec_polys = _get(ocr_res, "rec_polys") or []
        if rec_texts and rec_polys and len(rec_texts) == len(rec_polys):
            words = []
            for txt, poly in zip(rec_texts, rec_polys):
                txt = (txt or "").strip()
                if not txt:
                    continue
                try:
                    xs = [float(pt[0]) for pt in poly]
                    ys = [float(pt[1]) for pt in poly]
                except (TypeError, IndexError, ValueError):
                    continue
                x0, y0 = int(min(xs)), int(min(ys))
                x1, y1 = int(max(xs)), int(max(ys))
                words.append({
                    "text": txt,
                    "x": x0, "y": y0,
                    "w": max(1, x1 - x0),
                    "h": max(1, y1 - y0),
                })
            if words:
                rt = _build_table_from_words(words)
                if rt:
                    tables.append(rt)
                    logger.debug("paddle v3 p%d: OCR-only fallback (%d words → %d rows)",
                                 page_num, len(words), len(rt))

    return tables, titles


def _paddle_v2_extract(engine, arr, page_num: int) -> tuple[list[RichTable], list[str]]:
    """Legacy PPStructure (paddleocr 2.x) result schema."""
    try:
        result = engine(arr)
    except Exception as e:
        logger.warning("PPStructure p%d failed: %s", page_num, e)
        return [], []

    def _y0(r):
        b = r.get("bbox") or [0, 0, 0, 0]
        try:
            return float(b[1])
        except Exception:
            return 0.0
    try:
        result = sorted(result, key=_y0)
    except Exception:
        pass

    tables: list[RichTable] = []
    titles: list[str] = []
    for region in result:
        rtype = (region.get("type") or "").lower()
        res = region.get("res")
        if not res:
            continue

        if rtype == "table":
            html = res.get("html", "") if isinstance(res, dict) else ""
            rt = _parse_html_table(html)
            if rt:
                tables.append(rt)
                continue
            cells = res.get("cells") if isinstance(res, dict) else None
            if isinstance(cells, list) and cells:
                lines = [(c.get("text") or "").strip()
                         for c in cells if isinstance(c, dict)]
                lines = [ln for ln in lines if ln]
                if lines:
                    tables.append([[RichCell(text=ln)] for ln in lines])

        elif rtype == "title":
            pieces = []
            if isinstance(res, list):
                for item in res:
                    if isinstance(item, dict):
                        pieces.append((item.get("text") or "").strip())
                    elif isinstance(item, str):
                        pieces.append(item.strip())
            joined = " ".join(p for p in pieces if p)
            if joined:
                titles.append(joined)

        elif rtype in ("text", "figure_caption", "header", "footer"):
            lines = []
            if isinstance(res, list):
                for item in res:
                    if isinstance(item, dict):
                        t = (item.get("text") or "").strip()
                        if t:
                            lines.append(t)
            if lines:
                tables.append([[RichCell(text=ln)] for ln in lines])

    return tables, titles


# ---------- Extraction: RapidOCR (pip-only, broad Python support) ----------

def _get_rapidocr():
    """Lazily build a RapidOCR engine. Returns None if unavailable."""
    global _RAPIDOCR_ENGINE
    if _RAPIDOCR_ENGINE is False:
        return None
    if _RAPIDOCR_ENGINE is not None:
        return _RAPIDOCR_ENGINE
    if not _HAS_RAPIDOCR:
        _RAPIDOCR_ENGINE = False
        return None
    try:
        from rapidocr_onnxruntime import RapidOCR
        logger.info("Initializing RapidOCR (downloads models on first use)…")
        _RAPIDOCR_ENGINE = RapidOCR()
        logger.info("RapidOCR ready.")
        return _RAPIDOCR_ENGINE
    except Exception as e:
        logger.warning("RapidOCR init failed: %s", e)
        _RAPIDOCR_ENGINE = False
        return None


def _rapidocr_extract_page(pdf_path: str, page_num: int) -> tuple[list[RichTable], list[str]]:
    """OCR a page with RapidOCR, then reconstruct the table layout from the
    line-level bounding boxes via row/column clustering."""
    engine = _get_rapidocr()
    if engine is None:
        return [], []

    img = _render_page_image(pdf_path, page_num, dpi=300)
    if img is None:
        return [], []

    try:
        import numpy as np
        arr = np.array(img)
        if arr.ndim == 2:
            arr = np.stack([arr] * 3, axis=-1)
        elif arr.shape[-1] == 4:
            arr = arr[..., :3]
        result, _elapsed = engine(arr)
    except Exception as e:
        logger.warning("RapidOCR p%d failed: %s", page_num, e)
        return [], []

    if not result:
        return [], []

    # RapidOCR returns a list of [box, text, confidence] where box is 4 [x,y]
    # corner points (clockwise from top-left). Convert each line into a
    # word-like dict that _build_table_from_words can consume.
    words: list[dict] = []
    for entry in result:
        try:
            box = entry[0]
            text = (entry[1] or "").strip()
            conf = float(entry[2]) if len(entry) > 2 else 1.0
        except (IndexError, TypeError, ValueError):
            continue
        if not text or conf < 0.3:
            continue
        try:
            xs = [float(p[0]) for p in box]
            ys = [float(p[1]) for p in box]
        except (IndexError, TypeError, ValueError):
            continue
        x0, y0 = int(min(xs)), int(min(ys))
        x1, y1 = int(max(xs)), int(max(ys))
        words.append({
            "text": text,
            "x": x0,
            "y": y0,
            "w": max(1, x1 - x0),
            "h": max(1, y1 - y0),
        })

    if not words:
        return [], []

    rt = _build_table_from_words(words)
    return ([rt], []) if rt else ([], [])


# ---------- Extraction: smart Tesseract OCR (fallback) ----------

def _extract_ocr_page(pdf_path: str, page_num: int) -> Optional[RichTable]:
    """Tesseract OCR with row/column clustering — preserves table structure
    far better than dumping words line-by-line."""
    if not _HAS_TESSERACT:
        return None
    img = _render_page_image(pdf_path, page_num, dpi=400)
    if img is None:
        return None
    try:
        data = pytesseract.image_to_data(
            img,
            output_type=pytesseract.Output.DICT,
            config="--psm 6 -c preserve_interword_spaces=1",
        )
    except Exception as e:
        logger.warning("tesseract p%d failed: %s", page_num, e)
        return None

    words: list[dict] = []
    for i in range(len(data["text"])):
        txt = (data["text"][i] or "").strip()
        try:
            conf = int(data["conf"][i])
        except (ValueError, TypeError):
            conf = -1
        if not txt or conf < 30:
            continue
        words.append({
            "text": txt,
            "x": int(data["left"][i]),
            "y": int(data["top"][i]),
            "w": int(data["width"][i]),
            "h": int(data["height"][i]),
        })
    return _build_table_from_words(words)


def _build_table_from_words(words: list[dict]) -> Optional[RichTable]:
    """Cluster word boxes into rows by y-overlap, then assign each word to
    the column whose start it's nearest to (column starts found by 1-D
    clustering of word x-positions across all rows)."""
    if not words:
        return None

    # --- 1) cluster into rows by vertical overlap ---
    sorted_words = sorted(words, key=lambda w: w["y"])
    rows: list[list[dict]] = []
    for w in sorted_words:
        wt, wb = w["y"], w["y"] + w["h"]
        placed = False
        for row in rows:
            rt = min(x["y"] for x in row)
            rb = max(x["y"] + x["h"] for x in row)
            overlap = min(wb, rb) - max(wt, rt)
            if overlap > min(w["h"], rb - rt) * 0.4:
                row.append(w)
                placed = True
                break
        if not placed:
            rows.append([w])
    rows.sort(key=lambda r: sum(x["y"] for x in r) / len(r))
    rows = [sorted(r, key=lambda x: x["x"]) for r in rows]

    # --- 2) find column starts via 1-D clustering of x-positions ---
    median_w = sorted(w["w"] for w in words)[len(words) // 2]
    col_tol = max(15, int(median_w * 0.5))

    starts = sorted(w["x"] for w in words)
    clusters: list[list[int]] = []
    for x in starts:
        if not clusters or x - clusters[-1][-1] > col_tol:
            clusters.append([x])
        else:
            clusters[-1].append(x)

    # Drop very weak clusters (probably noise)
    n_total = max(1, len(starts))
    strong = [c for c in clusters if len(c) >= max(2, int(n_total * 0.04))]
    if not strong:
        strong = clusters
    col_starts = [c[0] for c in strong]

    if len(col_starts) <= 1:
        flat = [[RichCell(text=" ".join(w["text"] for w in row))] for row in rows]
        return _normalize_rich_table(flat)

    def col_of(x: int) -> int:
        ci = 0
        for i, cs in enumerate(col_starts):
            if x + col_tol // 2 >= cs:
                ci = i
            else:
                break
        return ci

    n_cols = len(col_starts)
    out: list[list[RichCell]] = []
    for row in rows:
        cells = [""] * n_cols
        for w in row:
            ci = col_of(w["x"])
            cells[ci] = (cells[ci] + " " + w["text"]).strip() if cells[ci] else w["text"]
        out.append([RichCell(text=c) for c in cells])
    return _normalize_rich_table(out)


# ---------- Pipeline ----------

def convert_pdf_to_workbook(pdf_path: str, pages: Optional[set[int]] = None) -> Workbook:
    """Backward-compatible wrapper. Returns just the Workbook."""
    wb, _summary = _convert_pdf_to_workbook_full(pdf_path, pages=pages)
    return wb


def _run_ocr_candidates(pdf_path: str, page_num: int) -> list:
    """Try every available OCR engine on this page. Returns scored candidates
    in the same shape as native extractors: (score, method, tables, titles)."""
    candidates: list[tuple[float, str, list[RichTable], list[str]]] = []

    # 1. PaddleOCR (PP-StructureV3 on 3.x, PPStructure on 2.x)
    p_tables, p_titles = _paddle_extract_page(pdf_path, page_num)
    if p_tables:
        sc = sum(_score_table(t) for t in p_tables) + 0.5
        tag = f"paddle-ocr-v{_PADDLE_API_VERSION}" if _PADDLE_API_VERSION else "paddle-ocr"
        candidates.append((sc, tag, p_tables, p_titles))
        logger.info("  paddle (%s): %d tables, score=%.2f", tag, len(p_tables), sc)
    elif _HAS_PADDLE:
        logger.info("  paddle: produced no tables")

    # 2. RapidOCR (ONNX-based)
    r_tables, r_titles = _rapidocr_extract_page(pdf_path, page_num)
    if r_tables:
        sc = sum(_score_table(t) for t in r_tables) + 0.4
        candidates.append((sc, "rapidocr", r_tables, r_titles))
        logger.info("  rapidocr: %d tables, score=%.2f", len(r_tables), sc)
    elif _HAS_RAPIDOCR:
        logger.info("  rapidocr: produced no tables")

    # 3. Tesseract — only if the others didn't fire (slow & weakest)
    if not p_tables and not r_tables and _HAS_TESSERACT:
        ocr_t = _extract_ocr_page(pdf_path, page_num)
        if ocr_t:
            sc = _score_table(ocr_t)
            candidates.append((sc, "tesseract-ocr", [ocr_t], []))
            logger.info("  tesseract: 1 table, score=%.2f", sc)

    return candidates


def _convert_pdf_to_workbook_full(
    pdf_path: str, pages: Optional[set[int]] = None,
    force_ocr: bool = False,
) -> tuple[Workbook, list[PageResult]]:
    """Build the workbook AND return the per-page method/score summary.

    When `force_ocr` is True, every selected page is OCR'd and the PDF's
    text layer is ignored entirely (use this when the text layer is garbage
    like ``(cid:N)`` glyph codes — copyable but meaningless)."""
    pp_results = _extract_pdfplumber(pdf_path, pages=pages)
    if pages is not None:
        page_count = max(pages) if pages else 0
    else:
        page_count = len(pp_results)

    # Skip camelot in force-OCR mode — its output is also from the text layer
    # we explicitly want to ignore.
    if force_ocr:
        lattice, stream = {}, {}
    else:
        lattice = _camelot_pages(pdf_path, "lattice", page_count, pages=pages)
        stream = _camelot_pages(pdf_path, "stream", page_count, pages=pages)

    chosen: list[PageResult] = []
    for r in pp_results:
        if force_ocr:
            logger.info("--- Page %d (FORCE OCR; native chars=%d ignored) ---",
                        r.page_num, r.char_count)
            candidates = _run_ocr_candidates(pdf_path, r.page_num)
            if not candidates:
                missing = []
                if not _HAS_PADDLE: missing.append("paddleocr")
                if not _HAS_RAPIDOCR: missing.append("rapidocr-onnxruntime")
                if not _HAS_TESSERACT: missing.append("tesseract")
                logger.warning("  no OCR engine produced output. Missing: %s",
                               ", ".join(missing) if missing else "(all engines failed)")
        else:
            logger.info("--- Page %d (chars=%d) ---", r.page_num, r.char_count)
            candidates = [(r.score, "pdfplumber", r.tables, r.title_lines)]
            logger.info("  pdfplumber: %d tables, score=%.2f",
                        len(r.tables), r.score)
            if r.page_num in lattice:
                ts = lattice[r.page_num]
                sc = sum(_score_table(t) for t in ts)
                candidates.append((sc, "camelot-lattice", ts, []))
                logger.info("  camelot/lattice: %d tables, score=%.2f", len(ts), sc)
            if r.page_num in stream:
                ts = stream[r.page_num]
                sc = sum(_score_table(t) for t in ts)
                candidates.append((sc, "camelot-stream", ts, []))
                logger.info("  camelot/stream:  %d tables, score=%.2f", len(ts), sc)

            # Auto-route to OCR when native is weak/empty
            best_native = max((c[0] for c in candidates), default=0.0)
            is_scanned = r.char_count < 5
            weak_native = best_native < 1.0 and sum(len(c[2]) for c in candidates) == 0
            if is_scanned or weak_native:
                why = "no text layer" if is_scanned else "native extractors empty"
                logger.info("  → OCR routing (%s)", why)
                candidates.extend(_run_ocr_candidates(pdf_path, r.page_num))

                if all(c[1] in ("pdfplumber", "camelot-lattice", "camelot-stream")
                       for c in candidates):
                    missing = []
                    if not _HAS_PADDLE: missing.append("paddleocr")
                    if not _HAS_RAPIDOCR: missing.append("rapidocr-onnxruntime")
                    if not _HAS_TESSERACT: missing.append("tesseract")
                    if missing:
                        logger.warning("  no OCR engine produced output. Missing: %s",
                                       ", ".join(missing))

        score, method, tables, titles = max(candidates, key=lambda c: c[0]) \
            if candidates else (0.0, "none", [], [])
        logger.info("  CHOSEN: %s (score %.2f, %d table%s)%s",
                    method, score, len(tables), "" if len(tables) == 1 else "s",
                    " [FORCE OCR]" if force_ocr else "")
        chosen.append(PageResult(
            page_num=r.page_num, tables=tables, title_lines=titles,
            method=method, score=score, char_count=r.char_count,
        ))

    return _build_workbook(chosen), chosen


# ---------- Excel rendering ----------

_THIN = Side(border_style="thin", color="FFB7B7B7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="FFE8EEF7", end_color="FFE8EEF7", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

_NUM_RE = re.compile(r"^\(?-?\$?\s*-?[\d,]+(\.\d+)?\)?\s*%?$")


def _looks_like_header_row(row: list[RichCell]) -> bool:
    non_empty = [c for c in row if c.text]
    if len(non_empty) < max(2, len(row) // 2):
        return False
    numeric = sum(1 for c in non_empty if _is_numeric(c.text))
    return numeric <= len(non_empty) // 3


def _is_numeric(s: str) -> bool:
    if not s:
        return False
    s2 = s.strip().replace(",", "").replace("$", "").replace("%", "")
    s2 = s2.replace("(", "-").replace(")", "")
    if not s2 or s2 in ("-", ".", "-."):
        return False
    try:
        float(s2)
        return True
    except ValueError:
        return False


def _parse_number(s: str) -> float:
    s2 = s.strip().replace(",", "").replace("$", "").replace("%", "")
    s2 = s2.replace("(", "-").replace(")", "")
    return float(s2)


def _is_section_row(text_row: list[str]) -> bool:
    """Row where only the first column has text (section/header divider)."""
    if not text_row or not text_row[0].strip():
        return False
    return all(not t.strip() for t in text_row[1:])


def _is_total_row(text_row: list[str]) -> bool:
    if not text_row or not text_row[0]:
        return False
    head = text_row[0].strip().lower()
    return (
        head.startswith("total ")
        or head.startswith("sub total")
        or head.startswith("subtotal")
        or head.startswith("net ")
        or head == "total"
        or head.startswith("grand total")
    )


def _autosize(ws, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            v = cell.value
            if v is None:
                continue
            for line in str(v).splitlines():
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def _write_title_rows(ws, title_lines: list[str], n_cols: int, start_row: int) -> int:
    if not title_lines:
        return start_row
    row = start_row
    for i, line in enumerate(title_lines):
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        size = 14 if i == 0 else (12 if i == 1 else 10)
        cell.font = Font(bold=(i <= 1), size=size, color="FF1F2937")
        if n_cols > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=n_cols)
        row += 1
    return row + 1  # spacer


def _apply_value_and_format(cell, raw: str) -> None:
    if _is_numeric(raw):
        try:
            num = _parse_number(raw)
        except ValueError:
            cell.value = raw
            return
        if "%" in raw:
            cell.value = num / 100
            cell.number_format = "0.00%;[Red]-0.00%"
        elif "$" in raw:
            cell.value = num
            cell.number_format = '_("$"* #,##0.00_);[Red]_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
        else:
            cell.value = num
            cell.number_format = "#,##0.00;[Red]-#,##0.00"
    else:
        cell.value = raw


def _write_rich_table(ws, table: RichTable, start_row: int, n_cols: int) -> int:
    if not table:
        return start_row

    is_header_row = [False] * len(table)
    is_header_row[0] = _looks_like_header_row(table[0])

    for r_off, row in enumerate(table):
        excel_row = start_row + r_off
        text_row = [c.text for c in row]
        section = _is_section_row(text_row)
        total = _is_total_row(text_row)
        header = is_header_row[r_off]

        for c_off, rich in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c_off)
            _apply_value_and_format(cell, rich.text)
            cell.border = _BORDER
            cell.alignment = Alignment(
                vertical="center",
                horizontal=("right" if _is_numeric(rich.text) else "left"),
                wrap_text=True,
            )

            font_kwargs = {}
            if rich.style.bold:
                font_kwargs["bold"] = True
            if rich.style.color_hex and rich.style.color_hex != "000000":
                font_kwargs["color"] = "FF" + rich.style.color_hex

            if header:
                font_kwargs["bold"] = True
                font_kwargs["color"] = "FFFFFFFF"
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(
                    vertical="center", horizontal="center", wrap_text=True
                )
            elif section:
                font_kwargs["bold"] = True
                font_kwargs["size"] = 11
                cell.fill = _SECTION_FILL
            elif total:
                font_kwargs["bold"] = True
                cell.fill = _TOTAL_FILL

            if font_kwargs:
                cell.font = Font(**font_kwargs)

        if section and n_cols > 1:
            try:
                ws.merge_cells(start_row=excel_row, start_column=1,
                               end_row=excel_row, end_column=n_cols)
            except Exception:
                pass  # already merged in some pathological case

    return start_row + len(table) + 1  # spacer row


def _build_workbook(pages: list[PageResult]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    any_content = False
    for page in pages:
        if not page.tables:
            continue
        any_content = True
        ws = wb.create_sheet(title=f"Page {page.page_num}"[:31])

        n_cols = max((len(t[0]) for t in page.tables if t), default=1)

        next_row = 1
        if page.title_lines:
            next_row = _write_title_rows(ws, page.title_lines, n_cols, next_row)

        for t in page.tables:
            next_row = _write_rich_table(ws, t, start_row=next_row, n_cols=n_cols)

        # Freeze panes below the title block (or at row 2 if no title)
        freeze_row = (1 + len(page.title_lines) + 1 + 1) if page.title_lines else 2
        ws.freeze_panes = f"A{freeze_row}"

        _autosize(ws, n_cols)

    if not any_content:
        ws = wb.create_sheet(title="No tables found")
        ws["A1"] = "No tables were detected in this PDF."
        engines = []
        if _HAS_PADDLE: engines.append("PaddleOCR")
        if _HAS_RAPIDOCR: engines.append("RapidOCR")
        if _HAS_TESSERACT: engines.append("Tesseract")
        if not engines:
            ws["A2"] = "No OCR engine is installed."
            ws["A3"] = "Run: pip install rapidocr-onnxruntime PyMuPDF"
        else:
            ws["A2"] = f"OCR engines available: {', '.join(engines)}"
            ws["A3"] = "But none could find structured content on these pages."
            if not _HAS_FITZ:
                ws["A4"] = "Tip: install PyMuPDF for reliable page rendering: pip install PyMuPDF"

    return wb


@dataclass
class ConvertResult:
    """Everything produced by one conversion run."""
    xlsx_bytes: bytes
    log_text: str
    ocr_used: bool
    ocr_engine: str  # which OCR was used (or "" if none)
    pages_summary: list[dict]  # [{page, method, score, tables, chars}, ...]
    preview_html: Optional[str] = None  # selectable-text preview of OCR'd content


def _engine_availability_block() -> str:
    """Produce the leading 'capability' block of the log file."""
    rows = [
        ("pdfplumber", True, "always (core)"),
        ("camelot",    _HAS_CAMELOT,    "tables w/ borders"),
        ("PyMuPDF",    _HAS_FITZ,       "page → image"),
        ("PaddleOCR",  _HAS_PADDLE,     "best OCR (table-aware)"),
        ("RapidOCR",   _HAS_RAPIDOCR,   "ONNX OCR (no paddle dep)"),
        ("Tesseract",  _HAS_TESSERACT,  "OCR fallback"),
    ]
    paddle_note = ""
    if _HAS_PADDLE:
        paddle_note = f" (api v{_PADDLE_API_VERSION})" if _PADDLE_API_VERSION else " (init pending)"
    out = ["=== boss-pdf · conversion log ===", ""]
    out.append("Available engines:")
    for name, ok, note in rows:
        flag = "✓" if ok else "✗"
        extra = paddle_note if name == "PaddleOCR" else ""
        out.append(f"  {flag} {name:<11} {note}{extra}")
    out.append("")
    return "\n".join(out)


def _render_preview_html(pages_summary: list[dict],
                         pages_tables: dict[int, list[RichTable]],
                         filename: str) -> str:
    """Render a standalone HTML file showing OCR'd content per page so the
    user can verify (selectable text → OCR worked)."""
    parts = [
        "<!DOCTYPE html>",
        "<html lang='en'><head><meta charset='utf-8'>",
        f"<title>boss-pdf · OCR preview · {_h(filename)}</title>",
        "<style>",
        "body{font-family:-apple-system,'Segoe UI',Inter,sans-serif;",
        "max-width:980px;margin:0 auto;padding:32px;background:#0d0e16;color:#f1eee7}",
        "h1{font-family:Georgia,serif;color:#c9a15a;margin:0 0 6px;font-size:30px}",
        ".lede{color:#9a9485;font-style:italic;margin:0 0 28px}",
        ".page{background:#171823;border:1px solid #2a2d3b;border-radius:14px;",
        "padding:20px 22px;margin:16px 0}",
        "h2{font-family:Georgia,serif;font-size:18px;margin:0 0 4px;color:#f1eee7}",
        ".meta{color:#7a7585;font-size:12px;margin-bottom:12px;letter-spacing:.3px}",
        "table{border-collapse:collapse;width:100%;font-size:13.5px;background:#fbfaf6;color:#0d0e16}",
        "td,th{border:1px solid #d8d3c4;padding:6px 8px;vertical-align:top;text-align:left}",
        "th{background:#e8dfc6}",
        "pre{white-space:pre-wrap;font-family:ui-monospace,Menlo,monospace;",
        "background:#0a0b12;border:1px solid #2a2d3b;color:#f1eee7;padding:14px;",
        "border-radius:8px;line-height:1.55;font-size:13px}",
        ".ok{color:#5fdfb0}.warn{color:#e0bd72}",
        "</style></head><body>",
        f"<h1>{_h(filename)} · OCR preview</h1>",
        "<p class='lede'>If the text below is <strong>readable and selectable</strong>, ",
        "OCR worked — you can use the .xlsx alongside this file. ",
        "If it looks garbled, check <code>boss-pdf-log.txt</code> in the same zip.</p>",
    ]
    for s in pages_summary:
        method = s["method"]
        if not method.endswith("ocr"):
            continue  # only show pages that actually went through OCR
        pn = s["page"]
        parts.append("<div class='page'>")
        parts.append(f"<h2>Page {pn}</h2>")
        parts.append(f"<div class='meta'>method: <strong>{_h(method)}</strong> · "
                     f"score: {s['score']:.2f} · "
                     f"native chars on page: {s['chars']}</div>")
        tables = pages_tables.get(pn) or []
        if not tables:
            parts.append("<pre class='warn'>(no content extracted)</pre>")
        for t in tables:
            parts.append(_richtable_to_html(t))
        parts.append("</div>")
    parts.append("</body></html>")
    return "\n".join(parts)


def _h(s: str) -> str:
    """Tiny HTML escape."""
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


def _richtable_to_html(t: RichTable) -> str:
    if not t:
        return ""
    out = ["<table>"]
    for ri, row in enumerate(t):
        out.append("<tr>")
        for c in row:
            tag = "th" if ri == 0 and any(cc.style.bold for cc in row) else "td"
            txt = _h(c.text or "").replace("\n", "<br>")
            style = ""
            if c.style.bold:
                style += "font-weight:700;"
            if c.style.color_hex and c.style.color_hex != "000000":
                style += f"color:#{c.style.color_hex};"
            attr = f" style='{style}'" if style else ""
            out.append(f"<{tag}{attr}>{txt}</{tag}>")
        out.append("</tr>")
    out.append("</table>")
    return "".join(out)


def convert_pdf_bytes_to_xlsx_bytes(pdf_bytes: bytes,
                                    pages: Optional[set[int]] = None) -> bytes:
    """Backward-compatible: returns just the xlsx bytes."""
    return convert_pdf_bytes(pdf_bytes, pages=pages, filename="").xlsx_bytes


def convert_pdf_bytes(pdf_bytes: bytes,
                      pages: Optional[set[int]] = None,
                      filename: str = "",
                      force_ocr: bool = False) -> ConvertResult:
    """Run the full pipeline and return ConvertResult: xlsx + log + preview.

    When ``force_ocr`` is True, the PDF's text layer (and Camelot, which also
    relies on it) is bypassed and every selected page is re-read with OCR.
    Useful for PDFs whose text layer is corrupted ``(cid:N)`` glyph codes."""
    # Capture every log message emitted from this module during the run.
    log_buf = io.StringIO()
    handler = logging.StreamHandler(log_buf)
    handler.setLevel(logging.DEBUG)
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s",
                                           datefmt="%H:%M:%S"))
    old_level = logger.level
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)

    log_buf.write(_engine_availability_block())
    logger.info("Input: %s (%d bytes)%s%s",
                filename or "<unnamed>", len(pdf_bytes),
                f", pages filter: {sorted(pages)}" if pages else "",
                " — FORCE OCR enabled" if force_ocr else "")

    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(pdf_bytes)
            tmp_path = f.name
        try:
            wb, page_results = _convert_pdf_to_workbook_full(
                tmp_path, pages=pages, force_ocr=force_ocr,
            )
            buf = io.BytesIO()
            wb.save(buf)
            xlsx_bytes = buf.getvalue()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    finally:
        logger.removeHandler(handler)
        logger.setLevel(old_level)

    pages_summary = [
        {"page": p.page_num, "method": p.method, "score": p.score,
         "tables": len(p.tables), "chars": p.char_count}
        for p in page_results
    ]
    ocr_used = force_ocr or any(p.method.endswith("ocr") for p in page_results)
    ocr_engine = next(
        (p.method for p in page_results if p.method.endswith("ocr")),
        ("forced-ocr-no-engine" if force_ocr else ""),
    )
    preview_html = None
    if ocr_used:
        pages_tables = {p.page_num: p.tables for p in page_results}
        preview_html = _render_preview_html(pages_summary, pages_tables,
                                            filename or "document.pdf")

    return ConvertResult(
        xlsx_bytes=xlsx_bytes,
        log_text=log_buf.getvalue(),
        ocr_used=ocr_used,
        ocr_engine=ocr_engine,
        pages_summary=pages_summary,
        preview_html=preview_html,
    )


# ---------- Page-range parsing ----------

def parse_page_spec(spec: str, total_pages: int) -> set[int]:
    """Parse '1-3,5,7-9' into {1,2,3,5,7,8,9}, clamped to [1, total_pages].

    Accepts 'all' (case-insensitive) or empty as 'every page'.
    Raises ValueError on malformed input or out-of-range pages.
    """
    if total_pages <= 0:
        raise ValueError("PDF has no pages.")
    if spec is None:
        return set(range(1, total_pages + 1))
    s = spec.strip().lower()
    if not s or s == "all":
        return set(range(1, total_pages + 1))

    out: set[int] = set()
    for chunk in s.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "-" in chunk:
            a, _, b = chunk.partition("-")
            try:
                start = int(a.strip())
                end = int(b.strip())
            except ValueError as e:
                raise ValueError(f"Invalid range '{chunk}'.") from e
            if start < 1 or end < 1 or start > end:
                raise ValueError(f"Invalid range '{chunk}'.")
            if end > total_pages:
                raise ValueError(
                    f"Range {chunk} exceeds page count ({total_pages})."
                )
            out.update(range(start, end + 1))
        else:
            try:
                p = int(chunk)
            except ValueError as e:
                raise ValueError(f"Invalid page '{chunk}'.") from e
            if p < 1 or p > total_pages:
                raise ValueError(
                    f"Page {p} out of range (1–{total_pages})."
                )
            out.add(p)
    if not out:
        raise ValueError("No pages selected.")
    return out
