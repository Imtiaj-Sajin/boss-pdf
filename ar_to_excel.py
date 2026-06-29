"""Convert a JDE EnterpriseOne 'A/R Details with Aging' (GGP0001) PDF into an
Excel workbook matching template1.xlsx.

The report is laid out in tenant blocks:

    Tenant: <id> <name>            Parent: <id> <name>
    Lease: <lease> Unit No : <unit> Type: <type> DBA: <id> <DBA-name> Status: ...
    Unit Type: ...
    <doc> <co> <G/L> <inv><due> <original> <open> <aging-amount> <remark...>
    ...
    Lease: <lease> <subtotal amounts>
    Total <name> <subtotal amounts>

Each charge line carries ONE aging amount; which aging bucket (Current / 1-30 /
31-60 / 61-90 / 91-120 / Over 120) it belongs to is decided by the amount's
horizontal position vs. the column headers — so this parser is coordinate-aware,
not text-only.

Output columns (B..R, matching template1.xlsx; column A left blank):
    Property Name | Tenant Name | Lease | Unit No | Type | G/L Offset |
    Invoice Date | Due Date/Check Date | Original Amount | Open Amount |
    Current | 1-30 | 31-60 | 61-90 | 91-120 | Over 120 | Remark

The tenant identity columns (B..F) are written only on the first charge row of
each tenant, exactly like the template.

Usage:
    python ar_to_excel.py INPUT.pdf [OUTPUT.xlsx] [--template template1.xlsx]
                                    [--totals] [--quiet]

  --no-totals  omit the per-block bold Total row (on by default)
  default OUTPUT is INPUT with a .xlsx extension, next to the PDF.

Run with no edits needed; it prints an aggregate reconciliation report
(no tenant data) comparing parsed sums to the report's own printed totals.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

NUM_RE = re.compile(r"^-?\(?[\d,]+\.\d{2}\)?-?$")
# Money as printed in this report: optional leading/paren minus, thousands
# separators, OPTIONAL decimals (some lines print whole dollars), optional
# trailing minus. Must contain a comma or a decimal point so bare integers
# (check numbers, etc.) aren't mistaken for amounts.
MONEY_RE = re.compile(r"^\(?-?(?:\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?|\d*\.\d{1,2})\)?-?$")
DATE_RE = re.compile(r"\d{1,2}/\d{1,2}/\d{4}")
DOC_RE = re.compile(r"^[A-Z]{2,3}\d{4,}$")            # e.g. DRD18319799
GLO_RE = re.compile(r"^[A-Z]{2,5}$")                  # e.g. BMRP, SLTX, UC (Unapplied Cash)

BUCKET_KEYS = ["Current", "1-30", "31-60", "61-90", "91-120", "Over120"]
HEADERS = ["Property Name", "Tenant Name", "Lease", "Unit No", "Type", "G/L Offset",
           "Invoice Date", "Due Date/Check Date", "Original Amount", "Open Amount",
           "Current", " 1 - 30", "31 - 60", "61 - 90", "91 - 120", "Over 120", "Remark"]
COL0 = 2                                              # first data column = B
BUCKET_COL = {"Current": 12, "1-30": 13, "31-60": 14, "61-90": 15, "91-120": 16, "Over120": 17}
MONEY = "#,##0.00"
DATEF = "mm-dd-yy"


# ----------------------------- parsing -----------------------------

def _num(s: str) -> Optional[float]:
    s = s.strip().replace(",", "")
    neg = False
    if s.endswith("-"):
        neg = True; s = s[:-1]
    if s.startswith("(") and s.endswith(")"):
        neg = True; s = s[1:-1]
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def _date(s: str) -> Optional[datetime]:
    try:
        return datetime.strptime(s, "%m/%d/%Y")
    except ValueError:
        return None


def _cx(w) -> float:
    return (w["x0"] + w["x1"]) / 2


@dataclass
class Charge:
    glo: str = ""
    inv: Optional[datetime] = None
    due: Optional[datetime] = None
    original: Optional[float] = None
    open: Optional[float] = None
    bucket_key: Optional[str] = None
    bucket_amt: Optional[float] = None
    remark: str = ""


@dataclass
class Tenant:
    property: Optional[str] = None
    tenant: Optional[str] = None
    lease: Optional[str] = None
    unit: Optional[str] = None
    type: Optional[str] = None
    charges: list = field(default_factory=list)
    printed_total: Optional[dict] = None     # the report's own "Total" line, for reconciliation


def _dedup_words(words, tol=3.0):
    """Drop overlapping duplicate glyphs (faux-bold / drop-shadow render, where
    the same text is painted 2-3x at a sub-pixel offset) so amounts aren't
    counted twice. Two words are the same glyph if identical text within `tol`
    px in BOTH axes. Genuine repeats (same value in another column, or the same
    charge on another row) sit far apart and are preserved."""
    kept = []
    for w in sorted(words, key=lambda w: (round(w["top"]), w["x0"])):
        if any(k["text"] == w["text"]
               and abs(k["x0"] - w["x0"]) < tol
               and abs(k["top"] - w["top"]) < tol
               for k in kept):
            continue
        kept.append(w)
    return kept


def _lines(page):
    words = _dedup_words(page.extract_words(x_tolerance=1.5, y_tolerance=2,
                                            keep_blank_chars=False))
    rows = defaultdict(list)
    for w in words:
        rows[round(w["top"])].append(w)
    return [sorted(rows[t], key=lambda w: w["x0"]) for t in sorted(rows)]


def _header_geometry(line):
    toks = [(w["text"], _cx(w)) for w in line]

    def find(txt, after=0):
        for i in range(after, len(toks)):
            if toks[i][0] == txt:
                return i
        return None

    def mid(i, j):
        return (toks[i][1] + toks[j][1]) / 2

    centers = {}
    i_cur = find("Current"); centers["Current"] = toks[i_cur][1]
    i1 = find("1", i_cur); centers["1-30"] = mid(i1, find("30", i1))
    i31 = find("31"); centers["31-60"] = mid(i31, find("60", i31))
    i61 = find("61"); centers["61-90"] = mid(i61, find("90", i61))
    i91 = find("91"); centers["91-120"] = mid(i91, find("120", i91))
    io = find("Over"); centers["Over120"] = mid(io, find("120", io))
    # Anchor the remark column on the LEFT edge (x0) of the 'Remark' header, not
    # its center — otherwise the threshold sits too far right and short
    # continuation words (e.g. a lone 'RENT') whose center falls left of it get
    # dropped, truncating multi-line remarks to their first line.
    remark_x = next((w["x0"] for w in line if w["text"] == "Remark"), 745.0)
    return centers, centers["Current"], remark_x


def _nearest_bucket(center, centers):
    return min(BUCKET_KEYS, key=lambda k: abs(center - centers[k]))


def _line_numbers(line, right_x=740.0):
    """Robustly extract (center_x, value) for every money figure on a line,
    stitching a detached trailing '-' (negative) onto the number it follows.
    Figures only occur between the Original column (x~205) and the remark
    column; `right_x` excludes numbers embedded in remark text (e.g. 'Emld 2.6')
    that would otherwise be mistaken for an aging amount."""
    out = []
    for w in sorted(line, key=lambda w: w["x0"]):
        c = _cx(w)
        if c < 205 or c >= right_x:
            continue
        t = w["text"]
        if t == "-" and out and (c - out[-1][0]) < 45:
            cc, v = out[-1]
            out[-1] = (cc, -abs(v))
            continue
        if MONEY_RE.match(t):
            v = _num(t)
            if v is not None:
                out.append((c, v))
    return out


def _amounts_by_x(line, current_x, centers, right_x=740.0):
    """Return (original, open, bucket_key, bucket_amt) from a line's money
    figures. A normal charge line has exactly three figures in order —
    Original, Open, and the single Aging amount — so classify positionally and
    snap the aging figure to its nearest bucket. (Positional beats x-bands here:
    wide negative numbers like '107,101.04-' shift their center left across the
    Open/Current boundary and would otherwise be misfiled.) Lines with a
    non-standard count fall back to x-band classification."""
    nums = _line_numbers(line, right_x)
    if len(nums) == 3:
        bc, bv = nums[2]
        return nums[0][1], nums[1][1], _nearest_bucket(bc, centers), bv
    original = open_amt = bucket_amt = None
    bucket_key = None
    for c, v in nums:
        if c < 279:
            original = v
        elif c < current_x - 5:
            open_amt = v
        else:
            bucket_key = _nearest_bucket(c, centers)
            bucket_amt = v
    return original, open_amt, bucket_key, bucket_amt


def parse(pdf_path):
    tenants: list[Tenant] = []
    cur: Optional[Tenant] = None
    cur_charge: Optional[Charge] = None
    as_of = None
    property_name = None
    centers = None
    current_x = 375.0
    remark_x = 744.0
    stop = False
    in_summary = False
    grand = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            if stop:
                break
            in_header = True
            for line in _lines(page):
                t = " ".join(w["text"] for w in line)
                first = line[0]["text"] if line else ""

                if "Report Summary" in t:
                    in_summary = True
                    continue
                if in_summary:
                    # In the trailing Report Summary, GL-offset rows start with a
                    # letter code; only the Grand Total amounts row starts with a
                    # number (its 'Grand Total' label sits on an adjacent line).
                    # That row holds 8 ordered figures: Original, Open, then the
                    # six aging buckets. Read them by order (positions shift for
                    # these wide numbers, so x-bands are unreliable here).
                    if NUM_RE.match(first):
                        vals = [_num(w["text"]) for w in line if NUM_RE.match(w["text"])]
                        vals = [v for v in vals if v is not None]
                        if len(vals) >= 8:
                            keys = ["original", "open"] + BUCKET_KEYS
                            grand = dict(zip(keys, vals[:8]))
                            stop = True
                            break
                    continue

                if as_of is None and t.startswith("As Of "):
                    m = DATE_RE.search(t)
                    if m:
                        as_of = m.group(0)
                if property_name is None and first == "Company" and len(line) >= 3:
                    seg = [w["text"] for w in line[2:]
                           if w["text"] not in ("Aging", "Date") and not DATE_RE.match(w["text"])]
                    property_name = " ".join(seg).strip() or None

                if in_header:
                    if "Remark" in t and "Current" in t:
                        centers, current_x, remark_x = _header_geometry(line)
                        in_header = False
                    continue

                # ---- structural lines ----
                if first == "Tenant:":
                    # One tenant can hold several leases. Each lease is its own
                    # block, created on its identity Lease: line below; the
                    # Tenant: line just resets context.
                    cur = None
                    cur_charge = None
                    continue

                if first == "Lease:" and any(w["text"] == "Unit" for w in line):
                    # Identity line — start a NEW block per lease so a tenant with
                    # multiple leases yields one row-group per lease (not just the
                    # last one).
                    cur = Tenant(property=property_name)
                    tenants.append(cur)
                    words = [w["text"] for w in line]
                    cur.lease = words[1] if len(words) > 1 else None
                    if "No" in words:
                        ni = words.index("No")
                        rest = words[ni + 1:]
                        if rest and rest[0] == ":":
                            rest = rest[1:]
                        cur.unit = rest[0] if rest else None
                    if "Type:" in words:
                        ti = words.index("Type:")
                        di = words.index("DBA:") if "DBA:" in words else len(words)
                        cur.type = " ".join(words[ti + 1:di]).strip()
                    if "DBA:" in words:
                        di = words.index("DBA:")
                        si = words.index("Status:") if "Status:" in words else len(words)
                        seg = words[di + 1:si]
                        if seg and seg[0].isdigit():
                            seg = seg[1:]
                        cur.tenant = " ".join(seg).strip()
                    cur_charge = None
                    continue

                if first == "Lease:":              # per-lease subtotal -> reconcile
                    # 'Lease: <n> <amounts>' (no Unit) is a lease's own subtotal.
                    # Attach it to that lease's block for reconciliation.
                    lease_no = line[1]["text"] if len(line) > 1 else None
                    o, k, _bk, _ba = _amounts_by_x(line, current_x, centers, remark_x - 6)
                    for blk in reversed(tenants):
                        if blk.lease == lease_no:
                            blk.printed_total = {"original": o, "open": k}
                            break
                    cur_charge = None
                    continue

                if first == "Total":               # tenant-level total (spans leases) -> skip
                    cur_charge = None
                    continue
                if first == "Unit" and len(line) > 1 and line[1]["text"] == "Type:":
                    continue

                # ---- remark continuation (only words in the remark column) ----
                if cur_charge is not None and line and min(_cx(w) for w in line) >= remark_x - 6:
                    cur_charge.remark = (cur_charge.remark + " "
                                         + " ".join(w["text"] for w in line)).strip()
                    continue

                # ---- detail charge line: a G/L offset code (x~102), at least one
                # money figure, AND invoice/due dates. (The document token is
                # unreliable — some types like 'DRU' render it split into
                # letters+digits — so we don't require it. Requiring a date is what
                # separates real charges from the trailing summary-section GL rows,
                # which carry amounts but no dates.) ----
                glo = next((w["text"] for w in line
                            if 95 <= _cx(w) <= 125 and GLO_RE.match(w["text"])), None)
                has_money = any(_cx(w) >= 205 and MONEY_RE.match(w["text"]) for w in line)
                date_text = "".join(w["text"] for w in line if 122 <= _cx(w) < 205)
                dts = DATE_RE.findall(date_text)
                if glo is None or not has_money or not dts or cur is None:
                    continue
                original, open_amt, bk, ba = _amounts_by_x(line, current_x, centers,
                                                           remark_x - 6)
                remark = " ".join(w["text"] for w in line if _cx(w) >= remark_x - 6).strip()

                cur_charge = Charge(
                    glo=glo,
                    inv=_date(dts[0]) if dts else None,
                    due=_date(dts[1]) if len(dts) > 1 else None,
                    original=original, open=open_amt,
                    bucket_key=bk, bucket_amt=ba, remark=remark,
                )
                cur.charges.append(cur_charge)

    return {"as_of": as_of, "property": property_name, "tenants": tenants, "grand": grand}


# ----------------------------- export -----------------------------

def _fmt_asof(s):
    d = _date(s) if s else None
    return f"{d.month}/{d.day}/{d.year}" if d else (s or "")


def export(data, out_path, template_path=None, totals=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(row=1, column=COL0, value=f"As of {_fmt_asof(data['as_of'])}")
    for j, h in enumerate(HEADERS):
        c = ws.cell(row=2, column=COL0 + j, value=h)
        c.font = Font(bold=True, size=10)

    r = 3
    for tn in data["tenants"]:
        if not tn.charges:
            continue
        first = True
        for ch in tn.charges:
            if first:
                lease = tn.lease
                ws.cell(row=r, column=2, value=tn.property)
                ws.cell(row=r, column=3, value=tn.tenant)
                ws.cell(row=r, column=4,
                        value=int(lease) if (lease and lease.isdigit()) else lease)
                ws.cell(row=r, column=5, value=tn.unit)
                ws.cell(row=r, column=6, value=tn.type)
                first = False
            ws.cell(row=r, column=7, value=ch.glo)
            if ch.inv:
                ws.cell(row=r, column=8, value=ch.inv).number_format = DATEF
            if ch.due:
                ws.cell(row=r, column=9, value=ch.due).number_format = DATEF
            if ch.original is not None:
                ws.cell(row=r, column=10, value=ch.original).number_format = MONEY
            if ch.open is not None:
                ws.cell(row=r, column=11, value=ch.open).number_format = MONEY
            if ch.bucket_key and ch.bucket_amt is not None:
                ws.cell(row=r, column=BUCKET_COL[ch.bucket_key],
                        value=ch.bucket_amt).number_format = MONEY
            ws.cell(row=r, column=18, value=ch.remark)
            r += 1

        if totals:
            # Per-block 'Total' row, mirroring the report's per-lease subtotal:
            # 'Total' label in the Due Date column, bold sums in Original, Open,
            # and each populated aging bucket.
            bold = Font(bold=True)
            ws.cell(row=r, column=9, value="Total").font = bold
            o = sum(c.original or 0 for c in tn.charges)
            k = sum(c.open or 0 for c in tn.charges)
            bk_sum = defaultdict(float)
            for c in tn.charges:
                if c.bucket_key and c.bucket_amt is not None:
                    bk_sum[c.bucket_key] += c.bucket_amt
            for col, val in ((10, o), (11, k)):
                cc = ws.cell(row=r, column=col, value=round(val, 2))
                cc.number_format = MONEY; cc.font = bold
            for key, val in bk_sum.items():
                cc = ws.cell(row=r, column=BUCKET_COL[key], value=round(val, 2))
                cc.number_format = MONEY; cc.font = bold
            r += 1

    ws.freeze_panes = "A3"
    if template_path and os.path.exists(template_path):
        try:
            t = load_workbook(template_path)["Sheet1"]
            for col, dim in t.column_dimensions.items():
                if dim.width:
                    ws.column_dimensions[col].width = dim.width
        except Exception:
            pass
    wb.save(out_path)
    return r - 3


# ------------------------- reconciliation (aggregate only) -------------------------

def reconcile(data):
    """Compare each tenant's parsed sums to the report's own printed Total line.
    Returns aggregate counts only — no tenant data."""
    ok = bad = no_printed = 0
    for tn in data["tenants"]:
        if not tn.charges:
            continue
        pt = tn.printed_total
        if not pt or pt.get("open") is None:
            no_printed += 1
            continue
        my_open = round(sum(c.open or 0 for c in tn.charges), 2)
        if abs(my_open - round(pt["open"], 2)) <= 0.02:
            ok += 1
        else:
            bad += 1
    return ok, bad, no_printed


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf")
    ap.add_argument("out", nargs="?")
    ap.add_argument("--template", default=os.path.join(os.path.dirname(__file__), "template1.xlsx"))
    ap.add_argument("--no-totals", dest="totals", action="store_false",
                    help="omit the per-block Total row")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args(argv)

    out = args.out or os.path.splitext(args.pdf)[0] + ".xlsx"
    data = parse(args.pdf)
    rows = export(data, out, template_path=args.template, totals=args.totals)

    tenants = [t for t in data["tenants"] if t.charges]
    charges = sum(len(t.charges) for t in tenants)
    ok, bad, no_printed = reconcile(data)
    # duplicate tenant blocks (same lease+unit appearing more than once)
    keys = defaultdict(int)
    for t in tenants:
        keys[(t.lease, t.unit)] += 1
    dup_blocks = sum(v - 1 for v in keys.values() if v > 1)

    if not args.quiet:
        print(f"output:            {out}")
        print(f"tenant blocks:     {len(tenants)}")
        print(f"charge rows:       {charges}")
        print(f"data rows written: {rows}")
        print(f"duplicate blocks (same lease+unit seen >1x): {dup_blocks}")
        print("--- reconciliation vs report's own per-tenant 'Total' lines ---")
        print(f"  tenants matching printed Total:   {ok}")
        print(f"  tenants NOT matching:             {bad}")
        print(f"  tenants with no printed Total:    {no_printed}")
        my = {"original": sum(c.original or 0 for t in tenants for c in t.charges),
              "open": sum(c.open or 0 for t in tenants for c in t.charges)}
        for key in BUCKET_KEYS:
            my[key] = sum(c.bucket_amt or 0 for t in tenants for c in t.charges
                          if c.bucket_key == key)
        g = data.get("grand")
        print("--- reconciliation vs report Grand Total (per column) ---")
        if g:
            print(f"  {'column':10} {'parsed':>16} {'report':>16} {'delta':>14}")
            allpass = True
            for key in ["original", "open"] + BUCKET_KEYS:
                pv = round(my[key], 2); rv = round(g.get(key, 0) or 0, 2)
                d = pv - rv
                if abs(d) >= 0.5:
                    allpass = False
                print(f"  {key:10} {pv:16,.2f} {rv:16,.2f} {d:+14,.2f}")
            print(f"  => {'PASS' if allpass else 'FAIL'}")
        else:
            print("  report Grand Total not found")
    return 0


if __name__ == "__main__":
    sys.exit(main())
